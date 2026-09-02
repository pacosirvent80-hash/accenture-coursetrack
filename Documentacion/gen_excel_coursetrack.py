"""
Genera CourseTrack_Export.xlsx a partir de un JSON exportado desde CourseTrack.
Uso: python gen_excel_coursetrack.py <ruta_json>
"""
import json
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Colores ────────────────────────────────────────────────────────────────
AZUL_HDR  = '1E3A5F'
AZUL_MID  = '2563EB'
GRIS_ROW  = 'F7F9FC'
BLANCO    = 'FFFFFF'

ESTADO_COLORES = {
    'entregado':     ('EEF2FF', '4338CA'),
    'en_revision':   ('F5F3FF', '7C3AED'),
    'incidencia':    ('FEF2F2', 'DC2626'),
    'en_correccion': ('FFF7ED', 'C2410C'),
    'aprobado':      ('F0FDF4', '15803D'),
    'publicado':     ('F0FDFA', '0F766E'),
    'en_pausa':      ('F8FAFC', '64748B'),
}
SEV_COLORES = {
    'Alta':  ('FEF2F2', 'DC2626'),
    'Media': ('FFF7ED', 'C2410C'),
    'Baja':  ('F0FDF4', '15803D'),
}

# ── Helpers ────────────────────────────────────────────────────────────────
def hdr_font():  return Font(bold=True, color=BLANCO, name='Segoe UI', size=10)
def hdr_fill():  return PatternFill('solid', fgColor=AZUL_HDR)
def cell_font(color='1E293B', bold=False):
    return Font(name='Segoe UI', size=9, color=color, bold=bold)
def thin_border():
    s = Side(style='thin', color='E2E8F0')
    return Border(left=s, right=s, top=s, bottom=s)
def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def left():   return Alignment(horizontal='left',   vertical='center', wrap_text=True)

def write_header(ws, cols):
    for ci, (header, width) in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=header)
        c.font      = hdr_font()
        c.fill      = hdr_fill()
        c.alignment = center()
        c.border    = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 22

def write_cell(ws, row, col, value, bold=False, color='1E293B',
               fill=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = cell_font(color=color, bold=bold)
    c.border    = thin_border()
    c.alignment = align or left()
    if fill:
        c.fill = PatternFill('solid', fgColor=fill)
    return c

def estado_badge(ws, row, col, estado):
    bg, fg = ESTADO_COLORES.get(estado, ('F8FAFC', '64748B'))
    label = estado.replace('_', ' ').capitalize()
    write_cell(ws, row, col, label, color=fg, fill=bg, align=center())

def sev_badge(ws, row, col, sev):
    bg, fg = SEV_COLORES.get(sev, ('F8FAFC', '64748B'))
    write_cell(ws, row, col, sev, color=fg, fill=bg, align=center())

def shade_row(ws, row, ncols, shade):
    if shade:
        for ci in range(1, ncols + 1):
            c = ws.cell(row=row, column=ci)
            if not c.fill or c.fill.fgColor.rgb in ('00000000', 'FFFFFFFF', BLANCO):
                c.fill = PatternFill('solid', fgColor=GRIS_ROW)

# ── Carga JSON ─────────────────────────────────────────────────────────────
json_path = sys.argv[1] if len(sys.argv) > 1 else 'coursetrack_export.json'
with open(json_path, encoding='utf-8') as f:
    courses = json.load(f)

wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════
#  HOJA 1 — CURSOS (resumen con estado actual)
# ══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Cursos'
ws1.freeze_panes = 'A2'

cols1 = [
    ('Nombre del curso',   38),
    ('Código',             14),
    ('Proveedor',          20),
    ('SCORM / Formato',    18),
    ('Idioma',             14),
    ('Duración',           10),
    ('Estado actual',      16),
    ('Versión actual',     13),
    ('Últ. modificación',  17),
    ('Notas',              40),
]
write_header(ws1, cols1)

for ri, c in enumerate(courses, 2):
    vers = c.get('versions', [])
    last = vers[-1] if vers else {}
    estado = last.get('status', '—')
    version = last.get('number', '—')
    shade = (ri % 2 == 0)

    write_cell(ws1, ri, 1, c.get('name',''),     bold=True)
    write_cell(ws1, ri, 2, c.get('code',''),     align=center())
    write_cell(ws1, ri, 3, c.get('provider',''))
    write_cell(ws1, ri, 4, c.get('scormVersion',''))
    write_cell(ws1, ri, 5, c.get('language',''),  align=center())
    write_cell(ws1, ri, 6, c.get('duration',''),  align=center())
    estado_badge(ws1, ri, 7, estado)
    write_cell(ws1, ri, 8, version,               align=center())
    write_cell(ws1, ri, 9, c.get('updatedAt',''), align=center())
    write_cell(ws1, ri,10, c.get('notes',''))
    shade_row(ws1, ri, 10, shade)
    ws1.row_dimensions[ri].height = 18

# ══════════════════════════════════════════════════════════════════════════
#  HOJA 2 — VERSIONES
# ══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Versiones')
ws2.freeze_panes = 'A2'

cols2 = [
    ('Curso',              36),
    ('Código',             13),
    ('Versión',            10),
    ('Estado',             16),
    ('Fecha',              12),
    ('Modificado por',     16),
    ('Incidencias',        12),
    ('Notas de versión',   40),
]
write_header(ws2, cols2)

ri = 2
for c in courses:
    for v in c.get('versions', []):
        shade = (ri % 2 == 0)
        ninc = len(v.get('incidents', []))
        write_cell(ws2, ri, 1, c.get('name',''),          bold=True)
        write_cell(ws2, ri, 2, c.get('code',''),          align=center())
        write_cell(ws2, ri, 3, v.get('number',''),        align=center())
        estado_badge(ws2, ri, 4, v.get('status',''))
        write_cell(ws2, ri, 5, v.get('date',''),          align=center())
        write_cell(ws2, ri, 6, v.get('modificadoPor','—'), align=center())
        write_cell(ws2, ri, 7, ninc if ninc else '—',
                   color='DC2626' if ninc else '15803D', align=center(), bold=ninc > 0)
        write_cell(ws2, ri, 8, v.get('notes',''))
        shade_row(ws2, ri, 8, shade)
        ws2.row_dimensions[ri].height = 18
        ri += 1

# ══════════════════════════════════════════════════════════════════════════
#  HOJA 3 — INCIDENCIAS
# ══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Incidencias')
ws3.freeze_panes = 'A2'

cols3 = [
    ('Curso',              30),
    ('Versión',            10),
    ('Categoría',          18),
    ('Severidad',          11),
    ('Título',             34),
    ('Descripción',        46),
    ('Fecha',              12),
    ('Estado',             12),
    ('Creada por',         14),
    ('Editada por',        14),
    ('Resuelta por',       14),
    ('Resolución',         36),
]
write_header(ws3, cols3)

ri = 2
for c in courses:
    for v in c.get('versions', []):
        for inc in v.get('incidents', []):
            shade = (ri % 2 == 0)
            resuelta = inc.get('resolved', False)
            estado_inc = 'Resuelta' if resuelta else 'Abierta'
            est_bg = 'F0FDF4' if resuelta else 'FEF2F2'
            est_fg = '15803D' if resuelta else 'DC2626'

            write_cell(ws3, ri,  1, c.get('name',''),           bold=True)
            write_cell(ws3, ri,  2, v.get('number',''),         align=center())
            write_cell(ws3, ri,  3, inc.get('category',''))
            sev_badge(ws3, ri,   4, inc.get('severity',''))
            write_cell(ws3, ri,  5, inc.get('title',''),        bold=True)
            write_cell(ws3, ri,  6, inc.get('description',''))
            write_cell(ws3, ri,  7, inc.get('date',''),         align=center())
            write_cell(ws3, ri,  8, estado_inc,
                       color=est_fg, fill=est_bg, align=center())
            write_cell(ws3, ri,  9, inc.get('creadaPor','—'),   align=center())
            write_cell(ws3, ri, 10, inc.get('editadoPor','—'),  align=center())
            write_cell(ws3, ri, 11, inc.get('resueltaPor','—'), align=center())
            write_cell(ws3, ri, 12, inc.get('resolution',''))
            shade_row(ws3, ri, 12, shade)
            ws3.row_dimensions[ri].height = 28
            ri += 1

# ── Guardar ────────────────────────────────────────────────────────────────
base = os.path.splitext(os.path.basename(json_path))[0]
out  = os.path.join(os.path.dirname(json_path), f'{base}_export.xlsx')
wb.save(out)
print(f'✓ Excel generado: {out}')
